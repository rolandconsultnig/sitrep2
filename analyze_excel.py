import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys

excel_file = "NPF_SITREP_Smart_Reporting_Template Draft.xlsx"

try:
    # Load the workbook to get sheet names
    wb = load_workbook(excel_file, data_only=True)
    sheet_names = wb.sheetnames
    
    print("=" * 80)
    print(f"EXCEL FILE ANALYSIS: {excel_file}")
    print("=" * 80)
    print(f"\nTotal Sheets: {len(sheet_names)}")
    print(f"Sheet Names: {', '.join(sheet_names)}")
    print("\n" + "=" * 80)
    
    # Analyze each sheet
    for sheet_name in sheet_names:
        print(f"\n{'=' * 80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'=' * 80}")
        
        # Read the sheet with pandas
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            
            print(f"\nDimensions: {df.shape[0]} rows × {df.shape[1]} columns")
            
            # Check for empty rows/columns
            non_empty_rows = df.dropna(how='all').shape[0]
            non_empty_cols = df.dropna(axis=1, how='all').shape[1]
            print(f"Non-empty rows: {non_empty_rows}")
            print(f"Non-empty columns: {non_empty_cols}")
            
            # Display first few rows
            print(f"\nFirst 20 rows (showing non-empty cells):")
            print("-" * 80)
            
            # Show a preview of the data
            preview_df = df.head(20)
            for idx, row in preview_df.iterrows():
                row_data = []
                for col_idx, val in enumerate(row):
                    if pd.notna(val):
                        row_data.append(f"Col{col_idx+1}: {str(val)[:50]}")
                if row_data:
                    print(f"Row {idx+1}: {' | '.join(row_data)}")
            
            # Check for formulas, merged cells, etc. using openpyxl
            ws = wb[sheet_name]
            
            # Check merged cells
            merged_ranges = list(ws.merged_cells.ranges)
            if merged_ranges:
                print(f"\nMerged Cells: {len(merged_ranges)} ranges")
                for i, merged_range in enumerate(merged_ranges[:10]):  # Show first 10
                    print(f"  {i+1}. {merged_range}")
                if len(merged_ranges) > 10:
                    print(f"  ... and {len(merged_ranges) - 10} more")
            
            # Check for formulas
            formula_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # formula
                        formula_count += 1
            
            if formula_count > 0:
                print(f"\nFormulas found: {formula_count}")
            
            # Check for data validation
            if ws.data_validations.dataValidation:
                print(f"\nData Validations: {len(ws.data_validations.dataValidation)}")
            
        except Exception as e:
            print(f"Error reading sheet with pandas: {e}")
            # Try with openpyxl directly
            ws = wb[sheet_name]
            print(f"\nSheet dimensions (openpyxl): {ws.max_row} rows × {ws.max_column} columns")
            
            # Show first few rows
            print("\nFirst 10 rows:")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                row_values = [str(val) if val is not None else '' for val in row]
                if any(row_values):
                    print(f"Row {row_idx}: {' | '.join([f'Col{i+1}: {v[:30]}' for i, v in enumerate(row_values) if v])}")
    
    wb.close()
    
except Exception as e:
    print(f"Error analyzing Excel file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
