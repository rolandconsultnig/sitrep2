from fastapi import FastAPI, Depends, HTTPException, status, WebSocket, WebSocketDisconnect, Request, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
from typing import List, Optional
import uvicorn
import uuid
import os
import openpyxl
from io import BytesIO
from fastapi.responses import StreamingResponse

from database import SessionLocal, engine, Base
from models import User, Submission, ReportConfig, Agency, Collaboration, AuditLog, AlertConfig
from schemas import (
    UserCreate, UserResponse, UserUpdate, UserProfileResponse, Token, 
    SubmissionCreate, SubmissionResponse,
    DailySitrepResponse, WeeklySitrepResponse, RiskMatrixResponse,
    MonthlySitrepResponse, QuarterlySitrepResponse, BiAnnualSitrepResponse,
    YearlySitrepResponse,
    ReportConfigUpdate, ReportConfigResponse,
    AlertConfigUpdate, AlertConfigResponse
)
from permissions import has_permission, can_modify_user, can_view_state, get_permissions, RANK_HIERARCHY, get_rank_level, can_access_rank
from auth import get_current_user, create_access_token, verify_password, get_password_hash
from report_generator import (
    generate_daily_sitrep, generate_weekly_sitrep,
    generate_daily_risk_matrix, generate_weekly_risk_matrix
)
from report_generator_extended import (
    generate_monthly_sitrep, generate_quarterly_sitrep,
    generate_bi_annual_sitrep, generate_yearly_sitrep
)
from audit_logger import log_action
from realtime import manager, notify_new_submission, notify_red_alert
from analytics import predict_hotspots, detect_anomalies, get_trend_forecast, get_performance_metrics
from pdf_generator import generate_daily_sitrep_pdf, generate_weekly_sitrep_pdf
import json
import base64
import logging

logger = logging.getLogger(__name__)

# Create database tables
Base.metadata.create_all(bind=engine)

# Initialize default admin user on startup
def init_default_admin():
    db = SessionLocal()
    try:
        admin = db.query(User).filter(User.username == "admin").first()
        if not admin:
            admin = User(
                username="admin",
                email="admin@npf.gov.ng",
                hashed_password=get_password_hash("admin123"),
                full_name="System Administrator",
                rank="IGP",
                service_number="NPF-00001",
                state="FCT",
                department="Force Headquarters",
                role="admin",
                is_active="true"
            )
            db.add(admin)
            db.commit()
            print("+ Default admin user created (admin/admin123)")
    except Exception as e:
        print(f"Admin init error: {e}")
    finally:
        db.close()

init_default_admin()

app = FastAPI(
    title="NPF Smart SITREP System",
    description="Nigeria Police Force Smart Situation Report System",
    version="1.0.0"
)

# CORS middleware - allow Netlify frontend and local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://earnest-starlight-95f193.netlify.app",
        "http://localhost:3000",
        "http://localhost:5173",
        "http://localhost:3010",
        "http://13.53.33.63:3010",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def root():
    return {"message": "NPF Smart SITREP System API", "version": "1.0.0", "status": "running"}

@app.get("/favicon.ico")
async def favicon():
    raise HTTPException(status_code=204)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# Dependency to get DB session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Authentication endpoints
@app.post("/token", response_model=Token)
async def login(
    form_data: OAuth2PasswordRequestForm = Depends(), 
    db: Session = Depends(get_db),
    request: Request = None
):
    user = db.query(User).filter(User.username == form_data.username).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Check if user is active
    if user.is_active != "true":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account is deactivated"
        )
    
    # Update last login
    user.last_login = datetime.utcnow()
    db.commit()
    
    # Audit log login (never block auth if audit table/commit fails in production)
    ip_address = request.client.host if request else None
    user_agent = request.headers.get("user-agent") if request else None
    try:
        log_action(
            db, user, "LOGIN", "USER",
            ip_address=ip_address,
            user_agent=user_agent,
            details={"login_time": datetime.utcnow().isoformat(), "rank": user.rank},
        )
    except Exception as e:
        logger.exception("LOGIN audit log failed: %s", e)
        try:
            db.rollback()
        except Exception:
            pass

    access_token = create_access_token(data={"sub": user.username, "state": user.state, "rank": user.rank})
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/api/users/register", response_model=UserResponse)
async def register_user(
    user: UserCreate, 
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    # Check permissions - only users with create_users permission can register
    if not has_permission(current_user, "create_users"):
        raise HTTPException(status_code=403, detail="You don't have permission to create users")
    
    # Check if user exists
    db_user = db.query(User).filter(User.username == user.username).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    
    # Check if service number exists
    existing_service = db.query(User).filter(User.service_number == user.service_number).first()
    if existing_service:
        raise HTTPException(status_code=400, detail="Service number already registered")
    
    # Check if user can create users of this rank
    if current_user.rank and user.rank:
        if not can_access_rank(current_user.rank, user.rank):
            raise HTTPException(status_code=403, detail="Cannot create user with rank higher than your own")
    
    # Create new user
    hashed_password = get_password_hash(user.password)
    db_user = User(
        username=user.username,
        email=user.email,
        hashed_password=hashed_password,
        full_name=user.full_name,
        rank=user.rank,
        service_number=user.service_number,
        state=user.state,
        lga=user.lga,
        department=user.department,
        phone_number=user.phone_number,
        role=user.role,
        is_active="true"
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    
    # Audit log
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "CREATE", "USER",
        resource_id=db_user.id,
        ip_address=ip_address,
        details={"created_user": user.username, "rank": user.rank}
    )
    
    return db_user

# Submission endpoints
@app.post("/api/submissions", response_model=SubmissionResponse)
async def create_submission(
    submission: SubmissionCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    # Generate unique submission ID
    submission_id = f"NPF-{uuid.uuid4().hex[:8].upper()}-{datetime.now().strftime('%Y%m%d')}"
    
    submission_dict = submission.dict()
    submission_dict['submission_id'] = submission_id
    submission_dict['submitted_by'] = current_user.id
    
    # Ensure state matches user's state if not admin
    if current_user.role != "admin":
        submission_dict['state'] = current_user.state
    
    db_submission = Submission(**submission_dict)
    db.add(db_submission)
    db.commit()
    db.refresh(db_submission)
    
    # Audit log
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "CREATE", "SUBMISSION",
        resource_id=db_submission.id,
        ip_address=ip_address,
        details={"submission_id": submission_id, "threat": submission.threat_domain, "severity": submission.severity}
    )
    
    # Trigger WebSocket notifications
    try:
        submission_data = {
            "id": db_submission.id,
            "submission_id": submission_id,
            "threat": submission.threat_domain,
            "state": db_submission.state,
            "severity": submission.severity
        }
        await manager.broadcast({
            "type": "new_submission",
            "data": submission_data,
            "timestamp": datetime.utcnow().isoformat()
        })
        
        # Check if this triggers a RED alert
        if submission.severity in ["High", "Critical"]:
            await notify_red_alert({
                "threat": submission.threat_domain,
                "state": db_submission.state,
                "severity": submission.severity,
                "submission_id": submission_id
            })
    except Exception as e:
        # Don't fail the submission if notification fails
        print(f"WebSocket notification error: {e}")
    
    return db_submission

@app.get("/api/submissions", response_model=List[SubmissionResponse])
async def get_submissions(
    state: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    threat_domain: Optional[str] = None,
    severity: Optional[str] = None,
    search: Optional[str] = None,
    limit: Optional[int] = None,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(Submission)
    
    # Filter by state if user is not admin
    if current_user.role != "admin":
        query = query.filter(Submission.state == current_user.state)
    elif state:
        query = query.filter(Submission.state == state)
    
    if start_date:
        query = query.filter(Submission.report_date >= start_date)
    if end_date:
        query = query.filter(Submission.report_date <= end_date)
    if threat_domain:
        query = query.filter(Submission.threat_domain == threat_domain)
    if severity:
        query = query.filter(Submission.severity == severity)
    if search:
        query = query.filter(
            (Submission.narrative.ilike(f"%{search}%")) |
            (Submission.lga_or_address.ilike(f"%{search}%"))
        )
    
    query = query.order_by(Submission.report_date.desc())
    
    if offset:
        query = query.offset(offset)
    if limit:
        query = query.limit(limit)
    
    return query.all()

@app.get("/api/submissions/{submission_id}", response_model=SubmissionResponse)
async def get_submission(
    submission_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    # Check access
    if current_user.role != "admin" and submission.state != current_user.state:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    return submission

@app.put("/api/submissions/{submission_id}", response_model=SubmissionResponse)
async def update_submission(
    submission_id: int,
    submission_update: SubmissionCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Update an existing submission"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    # Check access - only submitter or admin can update
    if current_user.role != "admin" and submission.submitted_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to update this submission")
    
    # Update fields
    update_data = submission_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(submission, key, value)
    
    db.commit()
    db.refresh(submission)
    
    # Audit log
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "UPDATE", "SUBMISSION",
        resource_id=submission_id,
        ip_address=ip_address,
        details={"updated_fields": list(update_data.keys())}
    )
    
    return submission

@app.post("/api/submissions/{submission_id}/attachments")
async def upload_attachment(
    submission_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Upload an attachment to a submission"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    # Check access
    if current_user.role != "admin" and submission.submitted_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Validate file type
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'video/mp4', 'video/quicktime', 
                     'application/pdf', 'application/msword', 
                     'application/vnd.openxmlformats-officedocument.wordprocessingml.document']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="File type not allowed")
    
    # Limit file size (10MB)
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    
    # Store as base64 (for simplicity - in production use cloud storage)
    file_data = {
        "filename": file.filename,
        "content_type": file.content_type,
        "size": len(contents),
        "data": base64.b64encode(contents).decode('utf-8'),
        "uploaded_at": datetime.utcnow().isoformat()
    }
    
    # Get existing attachments or create new list
    existing = json.loads(submission.attachments) if submission.attachments else []
    existing.append(file_data)
    submission.attachments = json.dumps(existing)
    
    db.commit()
    
    # Audit log
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "UPLOAD", "ATTACHMENT",
        resource_id=submission_id,
        ip_address=ip_address,
        details={"filename": file.filename, "size": len(contents)}
    )
    
    return {"message": "File uploaded successfully", "filename": file.filename}

@app.get("/api/submissions/{submission_id}/attachments")
async def get_attachments(
    submission_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get list of attachments for a submission"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    # Check access
    if current_user.role != "admin" and submission.state != current_user.state:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    attachments = json.loads(submission.attachments) if submission.attachments else []
    # Return metadata only, not the actual data
    return [{"filename": a["filename"], "content_type": a["content_type"], 
             "size": a["size"], "uploaded_at": a.get("uploaded_at")} for a in attachments]

@app.get("/api/submissions/{submission_id}/attachments/{index}")
async def download_attachment(
    submission_id: int,
    index: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download a specific attachment"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    # Check access
    if current_user.role != "admin" and submission.state != current_user.state:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    attachments = json.loads(submission.attachments) if submission.attachments else []
    if index >= len(attachments):
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    attachment = attachments[index]
    content = base64.b64decode(attachment["data"])
    
    return StreamingResponse(
        BytesIO(content),
        media_type=attachment["content_type"],
        headers={"Content-Disposition": f"attachment; filename={attachment['filename']}"}
    )

@app.delete("/api/submissions/{submission_id}")
async def delete_submission(
    submission_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Delete a submission"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    # Check access - only submitter or admin can delete
    if current_user.role != "admin" and submission.submitted_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this submission")
    
    # Audit log before deletion
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "DELETE", "SUBMISSION",
        resource_id=submission_id,
        ip_address=ip_address,
        details={"submission_id": submission.submission_id, "threat": submission.threat_domain}
    )
    
    db.delete(submission)
    db.commit()
    
    return {"message": "Submission deleted successfully"}

@app.post("/api/submissions/upload")
async def upload_submissions_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Upload Excel file with multiple submissions"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        
        # Try to find RAW_DATA sheet or use first sheet
        if 'RAW_DATA' in wb.sheetnames:
            ws = wb['RAW_DATA']
        else:
            ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Expected columns mapping
        column_map = {
            'Report_Date': 'report_date',
            'State': 'state',
            'LGA_or_Address': 'lga_or_address',
            'Threat_Domain': 'threat_domain',
            'Severity': 'severity',
            'Trend': 'trend',
            'Source_Reliability': 'source_reliability',
            'Source_Credibility': 'source_credibility',
            'Other_Agency': 'other_agency',
            'Narrative': 'narrative'
        }
        
        # Find column indices
        col_indices = {}
        for i, header in enumerate(headers):
            if header in column_map:
                col_indices[column_map[header]] = i
        
        # Process rows
        created_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            try:
                # Extract data from row
                report_date = row[col_indices.get('report_date', 1)] if 'report_date' in col_indices else None
                state = row[col_indices.get('state', 2)] if 'state' in col_indices else None
                
                # Skip if essential fields are missing
                if not state or not report_date:
                    continue
                
                # Check state access for non-admin users
                if current_user.role != 'admin' and state != current_user.state:
                    errors.append(f"Row {row_num}: Not authorized to submit for state {state}")
                    continue
                
                # Parse date
                if isinstance(report_date, datetime):
                    parsed_date = report_date
                elif isinstance(report_date, str):
                    try:
                        parsed_date = datetime.fromisoformat(report_date.replace('Z', '+00:00'))
                    except:
                        parsed_date = datetime.now()
                else:
                    parsed_date = datetime.now()
                
                # Create submission
                submission_id = f"SIT-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
                
                db_submission = Submission(
                    submission_id=submission_id,
                    report_date=parsed_date,
                    state=state,
                    lga_or_address=row[col_indices.get('lga_or_address', 3)] if 'lga_or_address' in col_indices else '',
                    threat_domain=row[col_indices.get('threat_domain', 4)] if 'threat_domain' in col_indices else '',
                    severity=row[col_indices.get('severity', 5)] if 'severity' in col_indices else 'Medium',
                    trend=row[col_indices.get('trend', 6)] if 'trend' in col_indices else None,
                    source_reliability=row[col_indices.get('source_reliability', 7)] if 'source_reliability' in col_indices else None,
                    source_credibility=row[col_indices.get('source_credibility', 8)] if 'source_credibility' in col_indices else None,
                    other_agency=row[col_indices.get('other_agency', 9)] if 'other_agency' in col_indices else None,
                    narrative=row[col_indices.get('narrative', 10)] if 'narrative' in col_indices else '',
                    submitted_by=current_user.id
                )
                
                db.add(db_submission)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        db.commit()
        
        # Audit log
        ip_address = request.client.host if request else None
        log_action(
            db, current_user, "UPLOAD", "SUBMISSIONS",
            ip_address=ip_address,
            details={"filename": file.filename, "created_count": created_count, "errors": len(errors)}
        )
        
        return {
            "message": f"Successfully uploaded {created_count} submissions",
            "created_count": created_count,
            "errors": errors[:10] if errors else []  # Return first 10 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to process file: {str(e)}")

# Report endpoints
@app.get("/api/reports/daily-sitrep", response_model=DailySitrepResponse)
async def get_daily_sitrep(
    report_date: Optional[datetime] = None,
    state: Optional[str] = None,
    dashboard_window_days: Optional[int] = Query(
        None,
        ge=1,
        le=90,
        description="If set (e.g. 14), aggregate incidents over this many days for dashboards instead of a single calendar day.",
    ),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not report_date:
        report_date = datetime.now().date()

    # Check permissions
    if state and not can_view_state(current_user, state):
        raise HTTPException(status_code=403, detail="Not authorized to view reports for this state")

    if dashboard_window_days is not None:
        sitrep = generate_daily_sitrep(
            db, report_date, current_user, lookback_days=dashboard_window_days
        )
    else:
        sitrep = generate_daily_sitrep(db, report_date, current_user)
    return sitrep

@app.get("/api/reports/weekly-sitrep", response_model=WeeklySitrepResponse)
async def get_weekly_sitrep(
    week_start: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not week_start:
        # Default to current week (Monday)
        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())
    
    sitrep = generate_weekly_sitrep(db, week_start, current_user)
    return sitrep

@app.get("/api/reports/daily-risk-matrix", response_model=List[RiskMatrixResponse])
async def get_daily_risk_matrix(
    report_date: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not report_date:
        report_date = datetime.now().date()
    
    matrix = generate_daily_risk_matrix(db, report_date, current_user)
    return matrix

@app.get("/api/reports/weekly-risk-matrix", response_model=List[RiskMatrixResponse])
async def get_weekly_risk_matrix(
    week_start: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not week_start:
        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())
    
    matrix = generate_weekly_risk_matrix(db, week_start, current_user)
    return matrix

# PDF Report endpoints
@app.get("/api/reports/daily-sitrep/pdf")
async def get_daily_sitrep_pdf(
    report_date: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate and download daily SITREP as PDF"""
    if not report_date:
        report_date = datetime.now().date()
    
    sitrep = generate_daily_sitrep(db, report_date, current_user)
    pdf_buffer = generate_daily_sitrep_pdf(sitrep)
    
    filename = f"Daily_SITREP_{report_date.strftime('%Y%m%d') if hasattr(report_date, 'strftime') else report_date}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/reports/weekly-sitrep/pdf")
async def get_weekly_sitrep_pdf(
    week_start: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate and download weekly SITREP as PDF"""
    if not week_start:
        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())
    
    sitrep = generate_weekly_sitrep(db, week_start, current_user)
    pdf_buffer = generate_weekly_sitrep_pdf(sitrep)
    
    filename = f"Weekly_SITREP_{week_start.strftime('%Y%m%d') if hasattr(week_start, 'strftime') else week_start}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Extended report endpoints
@app.get("/api/reports/monthly-sitrep", response_model=MonthlySitrepResponse)
async def get_monthly_sitrep(
    month_start: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not month_start:
        today = datetime.now()
        month_start = today.replace(day=1)
    
    sitrep = generate_monthly_sitrep(db, month_start, current_user)
    return sitrep

@app.get("/api/reports/quarterly-sitrep", response_model=QuarterlySitrepResponse)
async def get_quarterly_sitrep(
    quarter_start: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not quarter_start:
        today = datetime.now()
        quarter_start = today
    
    sitrep = generate_quarterly_sitrep(db, quarter_start, current_user)
    return sitrep

@app.get("/api/reports/bi-annual-sitrep", response_model=BiAnnualSitrepResponse)
async def get_bi_annual_sitrep(
    period_start: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not period_start:
        period_start = datetime.now()
    
    sitrep = generate_bi_annual_sitrep(db, period_start, current_user)
    return sitrep

@app.get("/api/reports/yearly-sitrep", response_model=YearlySitrepResponse)
async def get_yearly_sitrep(
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not year:
        year = datetime.now().year
    
    sitrep = generate_yearly_sitrep(db, year, current_user)
    return sitrep

# Configuration endpoints
@app.get("/api/config", response_model=ReportConfigResponse)
async def get_config(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    config = db.query(ReportConfig).first()
    if not config:
        # Create default config
        config = ReportConfig()
        db.add(config)
        db.commit()
        db.refresh(config)
    return config

@app.put("/api/config", response_model=ReportConfigResponse)
async def update_config(
    config_update: ReportConfigUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update config")
    
    config = db.query(ReportConfig).first()
    if not config:
        config = ReportConfig(**config_update.dict())
        db.add(config)
    else:
        for key, value in config_update.dict(exclude_unset=True).items():
            setattr(config, key, value)
    
    db.commit()
    db.refresh(config)
    return config

# Alert Configuration endpoints
@app.get("/api/alert-config")
async def get_alert_config(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get current user's alert configuration"""
    config = db.query(AlertConfig).filter(AlertConfig.user_id == current_user.id).first()
    if not config:
        # Return default config
        return {
            "email_enabled": "true",
            "sms_enabled": "false",
            "phone_number": current_user.phone_number,
            "alert_on_critical": "true",
            "alert_on_high": "true",
            "alert_on_medium": "false",
            "alert_on_low": "false",
            "alert_states": [current_user.state] if current_user.state else [],
            "daily_digest": "false",
            "digest_time": "08:00"
        }
    
    return {
        "id": config.id,
        "email_enabled": config.email_enabled,
        "sms_enabled": config.sms_enabled,
        "phone_number": config.phone_number,
        "alert_on_critical": config.alert_on_critical,
        "alert_on_high": config.alert_on_high,
        "alert_on_medium": config.alert_on_medium,
        "alert_on_low": config.alert_on_low,
        "alert_states": json.loads(config.alert_states) if config.alert_states else [],
        "daily_digest": config.daily_digest,
        "digest_time": config.digest_time
    }

@app.put("/api/alert-config")
async def update_alert_config(
    config_update: AlertConfigUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Update current user's alert configuration"""
    config = db.query(AlertConfig).filter(AlertConfig.user_id == current_user.id).first()
    
    update_data = config_update.dict(exclude_unset=True)
    
    # Convert alert_states list to JSON string
    if 'alert_states' in update_data and update_data['alert_states'] is not None:
        update_data['alert_states'] = json.dumps(update_data['alert_states'])
    
    if not config:
        config = AlertConfig(user_id=current_user.id, **update_data)
        db.add(config)
    else:
        for key, value in update_data.items():
            setattr(config, key, value)
    
    db.commit()
    db.refresh(config)
    
    # Audit log
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "UPDATE", "ALERT_CONFIG",
        resource_id=config.id,
        ip_address=ip_address,
        details={"updated_fields": list(update_data.keys())}
    )
    
    return {"message": "Alert configuration updated successfully"}

# User info endpoint
@app.get("/api/users/me", response_model=UserProfileResponse)
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    permissions = get_permissions(current_user.rank or "ASP")
    return UserProfileResponse(
        **current_user.__dict__,
        permissions=permissions
    )

# User management endpoints
@app.get("/api/users", response_model=List[UserResponse])
async def get_users(
    state: Optional[str] = None,
    rank: Optional[str] = None,
    is_active: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get list of users with filtering"""
    query = db.query(User)
    
    # Apply filters based on permissions
    if not has_permission(current_user, "view_all_states"):
        query = query.filter(User.state == current_user.state)
    elif state:
        query = query.filter(User.state == state)
    
    if rank:
        query = query.filter(User.rank == rank)
    
    if is_active:
        query = query.filter(User.is_active == is_active)
    
    # Users can only see users of equal or lower rank
    if current_user.rank:
        user_level = get_rank_level(current_user.rank)
        # Filter to show only users of equal or lower rank
        allowed_ranks = [r for r, level in RANK_HIERARCHY.items() if level <= user_level]
        query = query.filter(User.rank.in_(allowed_ranks))
    
    users = query.order_by(User.rank.desc(), User.full_name).all()
    return users

@app.get("/api/users/{user_id}", response_model=UserProfileResponse)
async def get_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get user profile by ID"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check if user can view this user
    if not has_permission(current_user, "view_all_states") and user.state != current_user.state:
        raise HTTPException(status_code=403, detail="Not authorized to view this user")
    
    # Check rank hierarchy
    if current_user.rank and user.rank:
        if not can_access_rank(current_user.rank, user.rank) and current_user.id != user_id:
            raise HTTPException(status_code=403, detail="Not authorized to view this user")
    
    permissions = get_permissions(user.rank or "ASP")
    return UserProfileResponse(
        **user.__dict__,
        permissions=permissions
    )

@app.put("/api/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user_update: UserUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Update user profile"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check permissions
    if not can_modify_user(current_user, user):
        raise HTTPException(status_code=403, detail="You don't have permission to modify this user")
    
    # Update fields
    update_data = user_update.dict(exclude_unset=True)
    
    # Prevent rank escalation beyond user's own rank
    if "rank" in update_data and update_data["rank"]:
        if not can_access_rank(current_user.rank or "ASP", update_data["rank"]):
            raise HTTPException(status_code=403, detail="Cannot assign rank higher than your own")
    
    for key, value in update_data.items():
        setattr(user, key, value)
    
    user.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(user)
    
    # Audit log
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "UPDATE", "USER",
        resource_id=user_id,
        ip_address=ip_address,
        details={"updated_fields": list(update_data.keys())}
    )
    
    return user

@app.delete("/api/users/{user_id}")
async def delete_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Delete/deactivate user"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Only users with delete permission can delete
    if not has_permission(current_user, "delete_users"):
        raise HTTPException(status_code=403, detail="You don't have permission to delete users")
    
    # Cannot delete self
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    # Soft delete - deactivate instead
    user.is_active = "false"
    user.updated_at = datetime.utcnow()
    db.commit()
    
    # Audit log
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "DELETE", "USER",
        resource_id=user_id,
        ip_address=ip_address,
        details={"deleted_user": user.username}
    )
    
    return {"message": "User deactivated successfully"}

@app.put("/api/users/{user_id}/activate")
async def activate_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Activate a deactivated user"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not can_modify_user(current_user, user):
        raise HTTPException(status_code=403, detail="You don't have permission to activate this user")
    
    user.is_active = "true"
    user.updated_at = datetime.utcnow()
    db.commit()
    
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "UPDATE", "USER",
        resource_id=user_id,
        ip_address=ip_address,
        details={"action": "activate"}
    )
    
    return {"message": "User activated successfully"}

# WebSocket endpoint for real-time updates
@app.websocket("/ws/{user_id}")
async def websocket_endpoint(websocket: WebSocket, user_id: int):
    await manager.connect(websocket, user_id)
    try:
        while True:
            data = await websocket.receive_text()
            # Echo back or process message
            await manager.send_personal_message({"message": "Received", "data": data}, websocket)
    except WebSocketDisconnect:
        manager.disconnect(websocket, user_id)

# Analytics endpoints
@app.get("/api/analytics/hotspots")
async def get_hotspot_predictions(
    days_ahead: int = 7,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get predicted crime hotspots"""
    predictions = predict_hotspots(db, days_ahead)
    log_action(db, current_user, "VIEW", "ANALYTICS", details={"type": "hotspots"})
    return {"predictions": predictions}

@app.get("/api/analytics/anomalies")
async def get_anomalies(
    state: Optional[str] = None,
    hours: int = 24,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Detect anomalies in recent data"""
    if current_user.role != "admin" and not state:
        state = current_user.state
    anomalies = detect_anomalies(db, state, hours)
    log_action(db, current_user, "VIEW", "ANALYTICS", details={"type": "anomalies"})
    return {"anomalies": anomalies}

@app.get("/api/analytics/forecast")
async def get_forecast(
    threat: str,
    state: Optional[str] = None,
    days: int = 30,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get trend forecast for a threat type"""
    if current_user.role != "admin" and not state:
        state = current_user.state
    forecast = get_trend_forecast(db, threat, state, days)
    log_action(db, current_user, "VIEW", "ANALYTICS", details={"type": "forecast", "threat": threat})
    return forecast

@app.get("/api/analytics/performance")
async def get_performance(
    days: int = 30,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get performance metrics"""
    metrics = get_performance_metrics(db, days)
    log_action(db, current_user, "VIEW", "ANALYTICS", details={"type": "performance"})
    return metrics

# Audit log endpoints
@app.get("/api/audit-logs")
async def get_audit_logs(
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get audit logs (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view audit logs")
    
    from audit_logger import get_audit_logs
    logs = get_audit_logs(db, user_id, action, resource_type, start_date, end_date, limit)
    return {"logs": [
        {
            "id": log.id,
            "username": log.username,
            "action": log.action,
            "resource_type": log.resource_type,
            "resource_id": log.resource_id,
            "ip_address": log.ip_address,
            "timestamp": log.timestamp.isoformat(),
            "details": log.details
        }
        for log in logs
    ]}

# Collaboration endpoints
@app.get("/api/agencies")
async def get_agencies(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get list of agencies"""
    agencies = db.query(Agency).filter(Agency.is_active == "true").all()
    return {"agencies": [
        {
            "id": a.id,
            "name": a.name,
            "code": a.code,
            "contact_email": a.contact_email,
            "contact_phone": a.contact_phone
        }
        for a in agencies
    ]}

@app.post("/api/collaborations")
async def share_submission(
    submission_id: int,
    agency_id: int,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Share a submission with another agency"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    # Check access
    if current_user.role != "admin" and submission.state != current_user.state:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    collaboration = Collaboration(
        submission_id=submission_id,
        agency_id=agency_id,
        shared_by=current_user.id,
        notes=notes
    )
    db.add(collaboration)
    db.commit()
    db.refresh(collaboration)
    
    log_action(
        db, current_user, "SHARE", "COLLABORATION",
        resource_id=collaboration.id,
        details={"submission_id": submission_id, "agency_id": agency_id}
    )
    
    return {"message": "Submission shared successfully", "collaboration_id": collaboration.id}

@app.get("/api/collaborations")
async def get_collaborations(
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get collaborations (shared submissions)"""
    query = db.query(Collaboration)
    
    if current_user.role != "admin":
        # Show collaborations where user shared or their state's submissions
        query = query.join(Submission).filter(
            (Collaboration.shared_by == current_user.id) |
            (Submission.state == current_user.state)
        )
    
    if status:
        query = query.filter(Collaboration.status == status)
    
    collaborations = query.all()
    return {"collaborations": [
        {
            "id": c.id,
            "submission_id": c.submission_id,
            "agency_id": c.agency_id,
            "agency_name": c.agency.name if c.agency else None,
            "shared_by": c.shared_by,
            "shared_at": c.shared_at.isoformat() if c.shared_at else None,
            "status": c.status,
            "notes": c.notes
        }
        for c in collaborations
    ]}

@app.put("/api/collaborations/{collaboration_id}/status")
async def update_collaboration_status(
    collaboration_id: int,
    status: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update collaboration status (accept/decline)"""
    if status not in ["pending", "accepted", "declined"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    collaboration = db.query(Collaboration).filter(Collaboration.id == collaboration_id).first()
    if not collaboration:
        raise HTTPException(status_code=404, detail="Collaboration not found")
    
    collaboration.status = status
    db.commit()
    
    log_action(
        db, current_user, "UPDATE", "COLLABORATION",
        resource_id=collaboration_id,
        details={"new_status": status}
    )
    
    return {"message": f"Collaboration {status}"}

# Agency management endpoints
@app.post("/api/agencies")
async def create_agency(
    name: str,
    code: str,
    contact_email: Optional[str] = None,
    contact_phone: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new agency"""
    if not has_permission(current_user, "manage_agencies"):
        raise HTTPException(status_code=403, detail="Not authorized to manage agencies")
    
    existing = db.query(Agency).filter((Agency.name == name) | (Agency.code == code)).first()
    if existing:
        raise HTTPException(status_code=400, detail="Agency with this name or code already exists")
    
    agency = Agency(
        name=name,
        code=code,
        contact_email=contact_email,
        contact_phone=contact_phone
    )
    db.add(agency)
    db.commit()
    db.refresh(agency)
    
    log_action(db, current_user, "CREATE", "AGENCY", resource_id=agency.id, details={"name": name})
    
    return {"message": "Agency created", "id": agency.id}

@app.put("/api/agencies/{agency_id}")
async def update_agency(
    agency_id: int,
    name: Optional[str] = None,
    code: Optional[str] = None,
    contact_email: Optional[str] = None,
    contact_phone: Optional[str] = None,
    is_active: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update an agency"""
    if not has_permission(current_user, "manage_agencies"):
        raise HTTPException(status_code=403, detail="Not authorized to manage agencies")
    
    agency = db.query(Agency).filter(Agency.id == agency_id).first()
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    
    if name:
        agency.name = name
    if code:
        agency.code = code
    if contact_email is not None:
        agency.contact_email = contact_email
    if contact_phone is not None:
        agency.contact_phone = contact_phone
    if is_active is not None:
        agency.is_active = is_active
    
    db.commit()
    
    log_action(db, current_user, "UPDATE", "AGENCY", resource_id=agency_id)
    
    return {"message": "Agency updated"}

@app.delete("/api/agencies/{agency_id}")
async def delete_agency(
    agency_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Deactivate an agency"""
    if not has_permission(current_user, "manage_agencies"):
        raise HTTPException(status_code=403, detail="Not authorized to manage agencies")
    
    agency = db.query(Agency).filter(Agency.id == agency_id).first()
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    
    agency.is_active = "false"
    db.commit()
    
    log_action(db, current_user, "DELETE", "AGENCY", resource_id=agency_id)
    
    return {"message": "Agency deactivated"}

# Logout endpoint
@app.post("/api/logout")
async def logout(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Logout and log the action"""
    ip_address = request.client.host if request else None
    user_agent = request.headers.get("user-agent") if request else None
    log_action(
        db, current_user, "LOGOUT", "USER",
        ip_address=ip_address,
        user_agent=user_agent,
        details={"logout_time": datetime.utcnow().isoformat()}
    )
    return {"message": "Logged out successfully"}

# Password change endpoint
@app.put("/api/users/me/password")
async def change_password(
    current_password: str,
    new_password: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None
):
    """Change current user's password"""
    if not verify_password(current_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    if len(new_password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    
    current_user.hashed_password = get_password_hash(new_password)
    current_user.updated_at = datetime.utcnow()
    db.commit()
    
    ip_address = request.client.host if request else None
    log_action(
        db, current_user, "UPDATE", "USER",
        resource_id=current_user.id,
        ip_address=ip_address,
        details={"action": "password_change"}
    )
    
    return {"message": "Password changed successfully"}

# Export endpoints
@app.get("/api/export/submissions")
async def export_submissions(
    format: str = "excel",
    state: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export submissions to Excel or CSV"""
    if not has_permission(current_user, "export_data"):
        raise HTTPException(status_code=403, detail="Not authorized to export data")
    
    from fastapi.responses import StreamingResponse
    import io
    import xlsxwriter
    
    query = db.query(Submission)
    
    if current_user.role != "admin":
        query = query.filter(Submission.state == current_user.state)
    elif state:
        query = query.filter(Submission.state == state)
    
    if start_date:
        query = query.filter(Submission.report_date >= start_date)
    if end_date:
        query = query.filter(Submission.report_date <= end_date)
    
    submissions = query.order_by(Submission.report_date.desc()).all()
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Submissions')
    
    # Headers
    headers = ['ID', 'Date', 'State', 'LGA/Address', 'Threat', 'Severity', 'Trend', 'Narrative']
    header_format = workbook.add_format({'bold': True, 'bg_color': '#0066cc', 'font_color': 'white'})
    
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # Data
    for row, sub in enumerate(submissions, 1):
        worksheet.write(row, 0, sub.submission_id)
        worksheet.write(row, 1, sub.report_date.strftime('%Y-%m-%d %H:%M') if sub.report_date else '')
        worksheet.write(row, 2, sub.state)
        worksheet.write(row, 3, sub.lga_or_address)
        worksheet.write(row, 4, sub.threat_domain)
        worksheet.write(row, 5, sub.severity)
        worksheet.write(row, 6, sub.trend or '')
        worksheet.write(row, 7, sub.narrative)
    
    worksheet.autofit()
    workbook.close()
    output.seek(0)
    
    log_action(db, current_user, "EXPORT", "SUBMISSION", details={"count": len(submissions), "format": format})
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=submissions_{datetime.now().strftime("%Y%m%d")}.xlsx'}
    )

@app.get("/api/export/report/{report_type}")
async def export_report(
    report_type: str,
    report_date: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export a report to Excel"""
    if not has_permission(current_user, "export_data"):
        raise HTTPException(status_code=403, detail="Not authorized to export data")
    
    from fastapi.responses import StreamingResponse
    import io
    import xlsxwriter
    
    if not report_date:
        report_date = datetime.now()
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    
    if report_type == "daily":
        sitrep = generate_daily_sitrep(db, report_date, current_user)
        worksheet = workbook.add_worksheet('Daily SITREP')
        
        title_format = workbook.add_format({'bold': True, 'font_size': 14})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#0066cc', 'font_color': 'white'})
        
        worksheet.write(0, 0, f'Daily SITREP - {report_date.strftime("%Y-%m-%d")}', title_format)
        worksheet.write(2, 0, 'Total Submissions (24h):', header_format)
        worksheet.write(2, 1, sitrep.total_submissions_24h)
        worksheet.write(3, 0, 'RED Alerts:', header_format)
        worksheet.write(3, 1, sitrep.red_alerts_count)
        
        # Incident Summary
        worksheet.write(5, 0, 'Incident Summary', title_format)
        worksheet.write(6, 0, 'Crime Type', header_format)
        worksheet.write(6, 1, 'Count', header_format)
        worksheet.write(6, 2, 'Severity', header_format)
        
        for row, incident in enumerate(sitrep.incident_summary, 7):
            worksheet.write(row, 0, incident.crime_type)
            worksheet.write(row, 1, incident.count)
            worksheet.write(row, 2, incident.severity_flag)
    
    workbook.close()
    output.seek(0)
    
    log_action(db, current_user, "EXPORT", "REPORT", details={"type": report_type})
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={report_type}_report_{datetime.now().strftime("%Y%m%d")}.xlsx'}
    )

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
